from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from selenium import webdriver
from selenium.webdriver.chrome.options import Options

import xml.etree.ElementTree as ET
import chromedriver_autoinstaller
from datetime import date

'''Check if the current version of chromedriver exists
and if it doesn't exist, download it automatically,
then add chromedriver to path'''
chromedriver_autoinstaller.install()

'# get today'
today_str = str(date.today())

# TODO need find out what abbreviations means, like SP, EP, DX and so on

'# Excel headers'
excel_headers = [  # Columns from url/det.asp
                 "Reading Date",
                 "Supply temperature",
                 "Extract temperature",
                 "Outdoor temperature",
                 "WT",
                 "Panel temperature 1",
                 "Panel temperature 2",
                 "Panel humidity RH 1",
                 "Panel humidity RH 2",
                 "Supply fan",
                 "Exhaust fan",
                 "SP",
                 "EP",
                 "SFI",
                 "EFI",
                 "S1",
                 "S2",
                 "Heat exchanger",
                 "WC",
                 "Electric heater",
                 "DX",
                 "Air dampers",
                 "Filter clogging",
                 "Energy saving",
                 "OH",
                 "Indoor humidity AH",

                 # Columns from url/i2.asp
                 "Ventilation level i2",
                 "Active hours i2",
                 "Next Away",

                 # Columns from url/i.asp
                 "Ventilation level i",
                 "Supply temperature",
                 "Extract temperature",
                 "Outdoor temperature",
                 "SP",
                 "SAF",
                 "EAF",
                 "SAFS",
                 "EAFS",
                 "Filter clogging",
                 "Heat exchanger efficiency",
                 "Heat recovery",
                 "Power consumption",
                 "Heating power",
                 "Specific power Acutal",
                 "Specific power Day",
                 "Consumed energy Day",
                 "Consumed energy Month",
                 "Consumed energy Total",
                 "Heating energy Day",
                 "Heating energy Month",
                 "Heating energy Total",
                 "Recovered energy Day",
                 "Recovered energy Month",
                 "Recovered energy Total",
                 "ST",
                 "ET",
                 "AQS",
                 "AQ",
                 "AHS",
                 "AH",
                 "VF"
                 ]  # all keys to be extracted from xml


def create_excel(excel_name, sheet_name):

    wb = Workbook()
    sheet1 = wb.active
    sheet1.title = sheet_name

    'freeze 1st row'
    sheet1.freeze_panes = 'A2'
    sheet1.append(excel_headers)

    '#add file type'
    file_name = excel_name + ".xlsx"
    '#salvestab exceli'
    wb.save(filename=file_name)


'# makes columns length wider'


def column_width(excel_name):
    '# add file type'
    file_name = excel_name + ".xlsx"

    workbook_name = file_name
    wb = load_workbook(workbook_name)
    sheet1 = wb.active

    '# 1 so enumerate, would start form 1, not 0'
    for i, col_value in enumerate(excel_headers, 1):
        '# if column length is very small (less then 5), then give static length of 10, else length of column'
        if len(col_value) < 5:
            column_extender = 10
        else:
            column_extender = len(col_value)
        '# wants column letter for input, as i. Width input is in the end of it'
        sheet1.column_dimensions[get_column_letter(i)].width = column_extender

    wb.save(filename=workbook_name)


'# You need to know your Komfovent Local IP aadress. det/i input need. As one is i.asp and other is det.asp'


def get_vent_stats(komfovent_local_ip, var):  # TODO PASSWORD from file and to gitignor it
    options = Options()
    '# parse without displaying Chrome'
    options.add_argument("--headless")
    '# UPDATE 9.01.2021 to avoid cannot find Chrome binary error'
    options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
    '# get Chrome driver with path'
    driver = webdriver.Chrome("chromedriver.exe", options=options)
    '# url we want to parse: Komfovent local IP'
    url = komfovent_local_ip
    '# get url'
    driver.get(url)

    '# If "CONTROL" are found on the top of the page then we are still logged in (previous session is still active) '
    'so skip the login part and go straight to API, else log in'
    if len(driver.find_elements_by_class_name('name')):
        pass
    else:
        driver.find_element_by_id('i_p').send_keys('user')
        driver.find_element_by_id('b_s').submit()

    '# get data from direct url API'
    driver.get(url + '//' + var + '.asp')#'//det.asp')
    '# get page source'
    data = driver.page_source
    '# Get XML part of the source'
    content = driver.find_element_by_id('webkit-xml-viewer-source-xml').get_attribute("innerHTML")
    '# Make page source string to XML'
    tree = ET.ElementTree(ET.fromstring(content))
    '# Get root of xml for the loop'
    root = tree.getroot()

    '# List to append the XML results'
    md_list = []
    for child in root:
        md_list.append(child.text.strip())

    return md_list


'# append new rows/info to excel'


def write_to_excel(excel_name, list_of_data):
    '# add file type'
    file_name = excel_name + ".xlsx"

    workbook_name = file_name
    wb = load_workbook(workbook_name)
    sheet1 = wb.active
    # New data to write:
    new_data = [list_of_data]

    for info in new_data:
        sheet1.append(info)

    wb.save(filename=workbook_name)


'# So insert would not return NONE'


def add_xpos_in_list(var, pos, iput_list):
    iput_list.insert(pos, var)

    return iput_list


data_combined_list = get_vent_stats("http://192.168.1.60/", 'det') + get_vent_stats("http://192.168.1.60/", 'i2') + get_vent_stats("http://192.168.1.60/", 'i')
new_vent_data_list = add_xpos_in_list(today_str, 0, data_combined_list)
#print(new_vent_data_list)

#create_excel('proov', 'sheet_name')
#column_width('proov', 'sheet_name' )
write_to_excel("proov", new_vent_data_list)


